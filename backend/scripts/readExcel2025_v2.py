"""
readExcel2025_v2.py - Parse 2025-26 CAP cutoff Excel files into JSON.
Handles TWO formats found in the 2025 data:

FORMAT A (CAP2/3/4 - combined cell):
  Data cell: "31722\n(83.80)"  → rank=31722, pct=83.80

FORMAT B (CAP1 - split rows):
  Row N:   [null, "11976", "40925"]  → ranks in row N (col 0 = null or label like Stage-I)
  Row N+1: [null, "-89.0", "-81.2"] → percentages in row N+1 (negative values = negative pct!)
  OR
  Row N:   ["Stage-I", "52319", "70776"]
  Row N+1: [null, "-78.4", "-74.0"]

Usage:
  python scripts/readExcel2025_v2.py <filename.xlsx>    ← process one file
  python scripts/readExcel2025_v2.py                    ← auto-process all known files
"""

import sys
import os
import json
from openpyxl import load_workbook

# Map known 2025-26 filenames → round number
FILE_TO_ROUND = {
    "POST_SSC_Diploma_CAP1_Cutoff.xlsx": 1,
    "POLY25_CAP_II_CUTOFF.xlsx": 2,
    "POLY_CAPIII_CUTOFF.xlsx": 3,
    "POLY_CAPIV_CUTOFF.xlsx": 4,
}

YEAR = 2025


def is_college_branch_header(row_vals):
    """Single-cell row: 'CODE - Name\\nCODE - Branch'"""
    if not row_vals:
        return False
    first = row_vals[0]
    if not first or not isinstance(first, str):
        return False
    rest = [c for c in row_vals[1:] if c not in (None, "")]
    if rest:
        return False  # multi-cell = not a header
    lines = first.strip().split("\n")
    return len(lines) >= 2 and " - " in lines[0] and " - " in lines[1]


def is_section_label(row_vals):
    """Single-cell text row (section label like 'Home District Seats')"""
    if not row_vals:
        return False
    first = row_vals[0]
    if not first or not isinstance(first, str):
        return False
    rest = [c for c in row_vals[1:] if c not in (None, "")]
    return len(rest) == 0 and "\n" not in first


def is_category_header_row(row_vals):
    """Row where col0 is None/empty and cols 1+ contain category strings."""
    if not row_vals:
        return False
    if row_vals[0] not in (None, "", 0):
        return False
    cats = [c for c in row_vals[1:] if c and isinstance(c, str) and c.strip()]
    return len(cats) > 0


def parse_combined_cell(val):
    """
    FORMAT A: "31722\n(83.80)"
    Returns (rank, percentage) or (None, None)
    """
    if not val:
        return None, None
    s = str(val).strip()
    if "(" not in s:
        return None, None
    parts = s.split("\n")
    try:
        rank = int(parts[0].strip())
    except (ValueError, IndexError):
        return None, None
    pct_part = next((p for p in parts if "(" in p), None)
    if not pct_part:
        return None, None
    try:
        pct = float(pct_part.replace("(", "").replace(")", "").strip())
    except ValueError:
        return None, None
    if rank != rank or pct != pct:  # NaN check
        return None, None
    return rank, pct


def is_format_a_value(val):
    """Check if a cell value is FORMAT A (combined rank+pct)."""
    if not val:
        return False
    s = str(val).strip()
    return "(" in s and "\n" in s


def is_numeric_value(val):
    """Check if val is a number (int or float)."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    try:
        float(str(val).strip())
        return True
    except ValueError:
        return False


def parse_data_sheet(rows):
    """
    Parse a data sheet (multi-row table).
    Auto-detects FORMAT A vs FORMAT B.
    Returns list of {category, rank, percentage}.
    """
    results = []

    # Find the category header row
    cat_row_idx = -1
    categories = []  # list of (col_idx, category_str)

    for r_idx, row in enumerate(rows):
        if not row:
            continue
        row_vals = list(row)

        if is_category_header_row(row_vals):
            cats = []
            for col_idx, val in enumerate(row_vals):
                if col_idx == 0:
                    continue
                if val and isinstance(val, str) and val.strip():
                    cats.append((col_idx, val.strip()))
            if cats:
                cat_row_idx = r_idx
                categories = cats
                break

    if cat_row_idx == -1 or not categories:
        return results

    # Peek at first data row to detect format
    data_rows = rows[cat_row_idx + 1:]

    # Detect format from first non-empty data row
    fmt = None
    for row in data_rows:
        if not row:
            continue
        row_vals = list(row)
        for col_idx, _ in categories:
            if col_idx < len(row_vals):
                val = row_vals[col_idx]
                if is_format_a_value(val):
                    fmt = "A"
                    break
                elif is_numeric_value(val):
                    fmt = "B"
                    break
        if fmt:
            break

    if fmt is None:
        return results

    if fmt == "A":
        # FORMAT A: single cell contains "rank\n(pct)"
        for row in data_rows:
            if not row:
                continue
            row_vals = list(row)
            for col_idx, category in categories:
                if col_idx >= len(row_vals):
                    continue
                rank, pct = parse_combined_cell(row_vals[col_idx])
                if rank is not None:
                    results.append({"category": category, "rank": rank, "percentage": pct})

    else:
        # FORMAT B: rank in one row, percentage in next row
        # Col 0 may be a stage label (Stage-I, Stage-II, etc.) or None
        i = 0
        data_rows_list = list(data_rows)
        while i < len(data_rows_list):
            row = data_rows_list[i]
            if not row:
                i += 1
                continue
            row_vals = list(row)

            # Check if this row has numeric values in category columns
            has_nums = any(
                col_idx < len(row_vals) and is_numeric_value(row_vals[col_idx])
                for col_idx, _ in categories
            )

            if has_nums:
                # This is a rank row — next row should be percentages
                rank_row = row_vals
                pct_row = []
                if i + 1 < len(data_rows_list) and data_rows_list[i + 1]:
                    pct_row = list(data_rows_list[i + 1])

                for col_idx, category in categories:
                    if col_idx >= len(rank_row):
                        continue
                    rank_val = rank_row[col_idx]
                    if not is_numeric_value(rank_val):
                        continue
                    try:
                        rank = int(float(str(rank_val).strip()))
                    except (ValueError, TypeError):
                        continue

                    pct = None
                    if pct_row and col_idx < len(pct_row):
                        pct_val = pct_row[col_idx]
                        if is_numeric_value(pct_val):
                            try:
                                pct = abs(float(str(pct_val).strip()))
                            except (ValueError, TypeError):
                                pass

                    if rank > 0:
                        results.append({"category": category, "rank": rank, "percentage": pct})

                # Skip the pct row
                if pct_row:
                    i += 2
                else:
                    i += 1
            else:
                i += 1

    return results


def parse_workbook(file_path, round_num):
    print(f"  Loading {os.path.basename(file_path)} ...", flush=True)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    records = []
    sheet_count = 0

    college_code = None
    college_name = None
    branch_code = None
    branch_name = None

    for sheet_name in wb.sheetnames:
        sheet_count += 1
        if sheet_count % 2000 == 0:
            print(f"    ... {sheet_count}/{len(wb.sheetnames)} sheets, {len(records)} records", flush=True)

        ws = wb[sheet_name]
        rows = list(ws.values)

        if not rows:
            continue

        row_vals = list(rows[0]) if rows[0] else []

        # Detect single-cell vs multi-cell sheet
        non_empty = [c for r in rows for c in (list(r) if r else []) if c not in (None, "")]
        cell_count = len(non_empty)

        if cell_count == 0:
            continue

        if cell_count == 1:
            # Single useful cell — header or section label
            val = str(row_vals[0]).strip() if row_vals else ""
            if " - " in val and "\n" in val:
                # College-branch header
                lines = val.split("\n")
                if len(lines) >= 2:
                    col_line = lines[0].strip()
                    bra_line = lines[1].strip()
                    if " - " in col_line and " - " in bra_line:
                        college_code = col_line.split(" - ", 1)[0].strip()
                        college_name = col_line.split(" - ", 1)[1].strip()
                        branch_code = bra_line.split(" - ", 1)[0].strip()
                        branch_name = bra_line.split(" - ", 1)[1].strip()
            # Section labels → no action
            continue

        # Multi-cell sheet = data table
        if not (college_code and branch_code):
            continue

        entries = parse_data_sheet(rows)
        for e in entries:
            records.append({
                "year": YEAR,
                "round": round_num,
                "collegeCode": college_code,
                "collegeName": college_name,
                "branchCode": branch_code,
                "branchName": branch_name,
                "category": e["category"],
                "rank": e["rank"],
                "percentage": e["percentage"],
            })

    wb.close()
    print(f"  Done: {sheet_count} sheets -> {len(records)} records", flush=True)
    return records


def main():
    cwd = os.getcwd()
    cli_file = sys.argv[1] if len(sys.argv) > 1 else None
    files_to_process = []

    if cli_file:
        file_path = cli_file if os.path.isabs(cli_file) else os.path.join(cwd, cli_file)
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            sys.exit(1)
        basename = os.path.basename(file_path)
        round_num = FILE_TO_ROUND.get(basename)
        if not round_num:
            print(f"Unknown file: {basename}. Known files: {list(FILE_TO_ROUND.keys())}")
            sys.exit(1)
        files_to_process = [(file_path, round_num)]
    else:
        for fname, rnd in FILE_TO_ROUND.items():
            fp = os.path.join(cwd, fname)
            if os.path.exists(fp):
                files_to_process.append((fp, rnd))
            else:
                print(f"Warning: not found: {fname}")

    if not files_to_process:
        print("No 2025-26 Excel files found.")
        sys.exit(1)

    for file_path, round_num in files_to_process:
        print(f"\n=== Processing round {round_num}: {os.path.basename(file_path)} ===", flush=True)
        records = parse_workbook(file_path, round_num)

        out_name = f"cutoffs2025_round{round_num}.json"
        out_path = os.path.join(cwd, out_name)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2)
        print(f"  Wrote {out_name} ({len(records)} records)")

        bad = [r for r in records if not all([r["collegeCode"], r["branchCode"], r["category"], r["rank"]])]
        print(f"  Bad records: {len(bad)}")

        seen = set()
        dups = 0
        for r in records:
            key = f"{r['collegeCode']}-{r['branchCode']}-{r['category']}-{r['rank']}"
            if key in seen:
                dups += 1
            seen.add(key)
        print(f"  Potential duplicates: {dups}")


if __name__ == "__main__":
    main()
