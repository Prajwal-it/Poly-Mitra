"""
inspectExcel2025.py - Inspect first N sheets of a 2025 workbook to understand structure.
Usage: python scripts/inspectExcel2025.py <filename.xlsx> [numSheets=10]
"""
import sys
import os
from openpyxl import load_workbook

file_path = sys.argv[1] if len(sys.argv) > 1 else None
num_sheets = int(sys.argv[2]) if len(sys.argv) > 2 else 10

if not file_path:
    print("Usage: python scripts/inspectExcel2025.py <filename.xlsx> [numSheets=10]")
    sys.exit(1)

if not os.path.isabs(file_path):
    file_path = os.path.join(os.getcwd(), file_path)

print(f"\nInspecting: {file_path}")
print(f"Showing first {num_sheets} sheets\n")

wb = load_workbook(file_path, read_only=True, data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"First {num_sheets} sheet names: {wb.sheetnames[:num_sheets]}\n")

for i, sheet_name in enumerate(wb.sheetnames[:num_sheets]):
    ws = wb[sheet_name]
    rows = list(ws.values)
    
    print("=" * 60)
    print(f"Sheet [{i}]: '{sheet_name}' ({len(rows)} rows)")
    
    if not rows:
        print("  (empty)")
        continue

    for r_idx, row in enumerate(rows[:5]):
        if row is None:
            print(f"  Row[{r_idx}]: None")
            continue
        row_vals = list(row)
        preview = []
        for c in row_vals[:8]:
            if c is None:
                preview.append("null")
            else:
                s = str(c).replace("\n", "\\n")[:60]
                preview.append(f'"{s}"')
        print("  Row[{}]: [{}]".format(r_idx, ", ".join(preview)))
    
    if len(rows) > 5:
        print(f"  ... ({len(rows) - 5} more rows)")

wb.close()
print("\nDone.")
