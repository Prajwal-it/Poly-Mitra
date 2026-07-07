"""
readExcel2025.py  –  Parse 2025-26 CAP cutoff Excel files into JSON.
Uses openpyxl with read_only + data_only mode for speed on large files.

Usage:
  python scripts/readExcel2025.py <filename.xlsx>
  python scripts/readExcel2025.py        # auto-discover all 4 known files
"""

import sys
import os
import re
import json
from openpyxl import load_workbook

# Map known 2025-26 filenames -> round number
FILE_TO_ROUND = {
    "POST_SSC_Diploma_CAP1_Cutoff.xlsx": 1,
    "POLY25_CAP_II_CUTOFF.xlsx": 2,
    "POLY_CAPIII_CUTOFF.xlsx": 3,
    "POLY_CAPIV_CUTOFF.xlsx": 4,
}

YEAR = 2025


def parse_rank_percentage(val):
    """
    Parse a cell value like '31722\n(83.80)' or '31722\n(83.80)\n...'
    Returns (rank, percentage) or (None, None).
    """
    if not val:
        return None, None
    s = str(val).strip()
    if "(" not in s:
        return None, None
    parts = s.split("\n")
    try:
        rank = int(parts[0].strip())
    except (ValueError, IndexError):
        return None, None
    pct_part = next((p for p in parts if "(" in p), None)
    if not pct_part:
        return None, None
    try:
        percentage = float(pct_part.replace("(", "").replace(")", "").strip())
    except ValueError:
        return None, None
    if rank != rank or percentage != percentage:  # NaN check
        return None, None
    return rank, percentage


def is_category_header_row(row_values):
    """Row where col0 is empty and at least one other col has a string."""
    if not row_values:
        return False
    if row_values[0] not in (None, "", 0):
        return False
    rest = [c for c in row_values[1:] if c and str(c).strip()]
    return len(rest) > 0


def is_section_label(row_values):
    """Row with a string in col0 and everything else empty."""
    if not row_values:
        return False
    if not isinstance(row_values[0], str):
        return False
    rest = [c for c in row_values[1:] if c not in (None, "", 0)]
    return len(rest) == 0


def parse_workbook(file_path, round_num):
    print(f"  Loading {os.path.basename(file_path)} ...", flush=True)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    records = []
    sheet_count = 0

    for sheet_name in wb.sheetnames:
        sheet_count += 1
        if sheet_count % 50 == 0:
            print(f"    ... processed {sheet_count} sheets, {len(records)} records so far", flush=True)

        ws = wb[sheet_name]
        rows = list(ws.values)

        if not rows or not rows[0] or not rows[0][0]:
            continue

        header_cell = str(rows[0][0]).strip()
        header_lines = header_cell.split("\n")

        if len(header_lines) < 2:
            continue

        college_line = header_lines[0].strip()
        branch_line = header_lines[1].strip()

        if " - " not in college_line or " - " not in branch_line:
            continue

        college_parts = college_line.split(" - ", 1)
        branch_parts = branch_line.split(" - ", 1)

        college_code = college_parts[0].strip()
        college_name = college_parts[1].strip()
        branch_code = branch_parts[0].strip()
        branch_name = branch_parts[1].strip()

        if not all([college_code, college_name, branch_code, branch_name]):
            continue

        current_categories = []  # list of (col_idx, category_str)

        for row in rows[1:]:
            if not row:
                continue
            row_vals = list(row)

            if is_section_label(row_vals):
                current_categories = []
                continue

            if is_category_header_row(row_vals):
                current_categories = []
                for col_idx, val in enumerate(row_vals[1:], start=1):
                    if val and str(val).strip():
                        current_categories.append((col_idx, str(val).strip()))
                continue

            if not current_categories:
                continue

            for col_idx, category in current_categories:
                if col_idx >= len(row_vals):
                    continue
                cell_val = row_vals[col_idx]
                rank, percentage = parse_rank_percentage(cell_val)
                if rank is None:
                    continue
                records.append({
                    "year": YEAR,
                    "round": round_num,
                    "collegeCode": college_code,
                    "collegeName": college_name,
                    "branchCode": branch_code,
                    "branchName": branch_name,
                    "category": category,
                    "rank": rank,
                    "percentage": percentage,
                })

    wb.close()
    print(f"  Done: {sheet_count} sheets → {len(records)} records", flush=True)
    return records


def main():
    cwd = os.getcwd()

    cli_file = sys.argv[1] if len(sys.argv) > 1 else None
    files_to_process = []

    if cli_file:
        file_path = cli_file if os.path.isabs(cli_file) else os.path.join(cwd, cli_file)
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            sys.exit(1)
        basename = os.path.basename(file_path)
        round_num = FILE_TO_ROUND.get(basename)
        if not round_num:
            print(f"Unknown file: {basename}. Known files: {list(FILE_TO_ROUND.keys())}")
            sys.exit(1)
        files_to_process = [(file_path, round_num)]
    else:
        for fname, rnd in FILE_TO_ROUND.items():
            fp = os.path.join(cwd, fname)
            if os.path.exists(fp):
                files_to_process.append((fp, rnd))
            else:
                print(f"Warning: not found: {fname}")

    if not files_to_process:
        print("No 2025-26 Excel files found.")
        sys.exit(1)

    for file_path, round_num in files_to_process:
        print(f"\nProcessing round {round_num}: {os.path.basename(file_path)}", flush=True)
        records = parse_workbook(file_path, round_num)

        out_name = f"cutoffs2025_round{round_num}.json"
        out_path = os.path.join(cwd, out_name)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(records, f, indent=2)
        print(f"  Wrote {out_name} ({len(records)} records)")

        bad = [r for r in records if not all([r["collegeCode"], r["branchCode"], r["category"], r["rank"], r["percentage"]])]
        print(f"  Bad records: {len(bad)}")

        seen = set()
        dups = 0
        for r in records:
            key = f"{r['collegeCode']}-{r['branchCode']}-{r['category']}"
            if key in seen:
                dups += 1
            seen.add(key)
        print(f"  Duplicates: {dups}")


if __name__ == "__main__":
    main()
